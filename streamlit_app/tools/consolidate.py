import streamlit as st

import openpyxl
from openpyxl import load_workbook
from openpyxl.cell.rich_text import TextBlock, CellRichText
from openpyxl.styles import PatternFill, Border, Side, Alignment, Protection, Font
from openpyxl.utils import get_column_letter
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.utils.cell import coordinate_from_string, column_index_from_string
from openpyxl.chart import BarChart, Reference

from fuzzywuzzy import fuzz
from itertools import cycle
from collections import deque, defaultdict
from copy import copy
import io
import re
import pandas as pd

from sumy.parsers.plaintext import PlaintextParser
from sumy.nlp.tokenizers import Tokenizer
from sumy.summarizers.lsa import LsaSummarizer


def fill_color_switch():
    """
    Returns a cycle of 10 colors used to fill cells in a worksheet. Each color represents a different supplier.

    The colors are chosen to be visually distinct and not too bright (color blindness friendly).
    """
    colors = [
        "E4DFEC",
        "D4D5F8",
        "DFD2FA",
        "D9E5F3",
        "E0E0EC",
        "E7DAF2",
        "E3E9E9",
        "CC79A7",
        "009E73",
        "0072B2",
    ]
    return cycle(colors)


def generate_merged_dict(sheet):
    """
    Generates a dictionary mapping column indices to lists of merged ranges
    in that column.

    The output dictionary will have the following structure:
    {
        column_idx1: [merged_range1, merged_range2, ...],
        column_idx2: [merged_range3, merged_range4, ...],
        ...
    }

    Only vertical merges are handled (i.e., merged cells in the same column).
    """
    merged_dict = {}

    for merged_range in sheet.merged_cells.ranges:
        # Only handle vertical merges (i.e., merged cells in the same column)
        if merged_range.size["columns"] == 1:
            start_cell = merged_range.start_cell
            column_idx = start_cell.column  # Get the column index
            merged_range_str = (
                merged_range.coord
            )  # Get the range as string (e.g., 'C14:C17')

            # If the column is not already in merged_dict, initialize an empty list
            if column_idx not in merged_dict:
                merged_dict[column_idx] = []

            # Add the merged range to the list of merged ranges for this column
            merged_dict[column_idx].append(merged_range_str)

    return merged_dict


def merge_columns_in_target_sheet(
    target_sheet, merged_dict, source_col_idx, target_col_idx
):
    """
    Merges columns in the target sheet based on the merged ranges found in the source sheet.

    Args:
        target_sheet: The target sheet where the columns will be merged.
        merged_dict: A dictionary mapping column indices to lists of merged ranges.
        source_col_idx: The column index from the source sheet where the merged ranges are found.
        target_col_idx: The column index in the target sheet where the columns will be merged.

    Returns:
        None
    """
    if source_col_idx not in merged_dict:
        # print(f"No merged ranges found for column {source_col_idx} in merged_dict.")
        return
    for merged_range in merged_dict[source_col_idx]:
        # Extract the start and end rows from the merged range (e.g., 'C14:C17' -> 14, 17)
        start_cell, end_cell = merged_range.split(":")
        start_row = int(start_cell[1:])
        end_row = int(end_cell[1:])
        start_column = column_index_from_string(start_cell[0])
        target_sheet.merge_cells(
            start_row=start_row,
            start_column=target_col_idx,  # Use target_col_idx for the target column
            end_row=end_row,
            end_column=target_col_idx,
        )


# Function to find common columns by comparing values
def find_matching_cols(template_sheet, supplier_sheet, threshold=80):
    """
    Find the columns that are common between the template sheet and the supplier sheet.

    Args:
        template_sheet: The template sheet.
        supplier_sheet: The supplier sheet.
        threshold: The threshold for fuzzy matching.

    Returns:
        A tuple of three lists: common_columns, mis_mat_rows, and supplier_value_columns.
            common_columns: The columns that are common between the template and the supplier.
            mis_mat_rows: The rows that have mismatched values between the template and the supplier.
            supplier_value_columns: The columns that contain the supplier values.
    """
    common_columns = []
    supplier_value_columns = []
    mis_mat_rows = []

    for col in template_sheet.iter_cols(max_col=min(template_sheet.max_column, 100)):
        # Get the column letter of the current column
        col_letter = get_column_letter(col[0].column)
        # Get the values of the current column in the template sheet
        row_values_template = []
        for cell in col:
            row_value = cell.value
            if row_value is not None:
                # If the cell is rich text, convert to string
                if isinstance(row_value, CellRichText):
                    row_values_template.append(" ".join(row_value.as_list()))
                else:
                    row_values_template.append(str(row_value))
        # Get the values of the current column in the supplier sheet
        row_values_suppliers = []
        for row_sup in supplier_sheet.iter_rows(
            min_col=col[0].column,
            max_col=col[0].column,
            max_row=min(supplier_sheet.max_row, 300),
        ):
            for cell in row_sup:
                if cell.value is not None:
                    # If the cell is rich text, convert to string
                    if isinstance(cell.value, CellRichText):
                        row_values_suppliers.append(" ".join(cell.value.as_list()))
                    else:
                        row_values_suppliers.append(str(cell.value))
        # If both lists are empty, skip the column
        if not row_values_template and not row_values_suppliers:
            continue

        # Fuzzy matching between the string joint from the template and the supplier list
        temp_row_str = " ".join(row_values_template)
        suppliers_row_str = " ".join(row_values_suppliers)
        similarity = fuzz.ratio(temp_row_str, suppliers_row_str)

        if similarity > threshold:
            common_columns.append(col_letter)
            if similarity < 100:
                # Highlight the row in the supplier sheet that does not match the template
                for row_sup in supplier_sheet.iter_rows(
                    min_col=col[0].column,
                    max_col=col[0].column,
                    max_row=min(supplier_sheet.max_row, 300),
                ):
                    for cell in row_sup:
                        cell_val, cell_coord = cell.value, cell.coordinate
                        if cell_val is not None:
                            # Check if the cell is rich text
                            if isinstance(cell_val, CellRichText):
                                cell_val = " ".join(cell_val.as_list())
                            else:
                                cell_val = str(cell_val)
                            if cell_val not in row_values_template:
                                print(
                                    f"Detected mismatch in row: {cell_coord} for supplier {supplier_sheet.title}. the value is: {cell_val}, the type is: {type(cell_val)}"
                                )
                                mis_mat_rows.append(cell_coord)
                                # Detected mismatch in row, add coordinates of the mismatched row

        else:
            # This column is not common, can be the column that contains the supplier values
            if row_values_suppliers:
                supplier_value_columns.append(col_letter)

    return common_columns, mis_mat_rows, supplier_value_columns


def separate_sheet_combine(
    workbook, template_sheets, supplier_sheets_dict, threshold=80
):
    """
    Combine files side by side in the same sheet.

    Parameters:
    workbook (openpyxl.Workbook): The workbook to combine the sheets into.
    template_sheets (list): A list of template sheets to combine.
    supplier_sheets_dict (dict): A dictionary mapping supplier names to their
        corresponding sheets to combine.
    threshold (int): The threshold for fuzzy matching between the template and
        supplier sheets.

    Returns:
    openpyxl.Workbook: The combined workbook with all the sheets.
    """
    st.toast(f"Combining files in progress...", icon="⏳")
    for idx, template_sheet in enumerate(template_sheets):
        # copy the template sheet to the workbook
        target_sheet = workbook.create_sheet(f"{template_sheet.title}"[:30])
        copy_sheet(template_sheet, target_sheet)
        for supplier in supplier_sheets_dict:
            print(f"Processing supplier: {supplier}")
            # add a new sheet for each supplier
            supplier_sheet = supplier_sheets_dict[supplier][idx]
            sheet_title = f"{supplier} {template_sheet.title}"[:30]
            target_sheet = workbook.create_sheet(sheet_title)
            copy_sheet(supplier_sheet, target_sheet)
            # Find matching columns between the template and supplier sheets
            _, mis_mat_rows, _ = find_matching_cols(
                template_sheet, supplier_sheet, threshold
            )
            # Highlight the mismatched rows
            if mis_mat_rows:
                for mis_mat_row in mis_mat_rows:
                    col_mis, row_mis = openpyxl.utils.cell.coordinate_from_string(
                        mis_mat_row
                    )
                    row_mis = int(row_mis)  # Convert to integer
                    col_mis = column_index_from_string(col_mis)  # Convert to integer
                    if col_mis and row_mis:
                        target_sheet.cell(row=row_mis, column=col_mis).fill = (
                            PatternFill(
                                start_color="FAA0A0",
                                end_color="FAA0A0",
                                fill_type="solid",
                            )
                        )
    # remove the default sheet
    if "Sheet" in workbook.sheetnames:
        workbook.remove(workbook["Sheet"])
    st.toast("Seperate-Sheet File combined successfully! Ready to download", icon="🎉")
    return workbook


def create_insertion_queue(common_columns, supplier_value_columns_dict):
    """
    Create a queue of columns to be inserted into the final combined sheet.

    The queue is sorted by column letter and includes a secondary sort by source type.
    Priority: template columns come before supplier columns with the same letter.

    Parameters:
    common_columns (list): A list of common column letters between the template and supplier sheets.
    supplier_value_columns_dict (dict): A dictionary mapping supplier names to their corresponding
        columns of values to be inserted.

    Returns:
    deque: The insertion queue.
    """
    # Initialize a list to hold queue items
    queue_items = []

    # Add common columns from the template to the queue
    for col in common_columns:
        queue_items.append({"column_letter": col, "source": "template"})

    # Add supplier columns to the queue
    for supplier, supplier_columns in supplier_value_columns_dict.items():
        for col in supplier_columns:
            queue_items.append({"column_letter": col, "source": supplier})

    # Sort the queue by column letter and include a secondary sort by source type
    # Priority: template columns come before supplier columns with the same letter
    queue_items.sort(key=lambda x: (x["column_letter"], x["source"] != "template"))

    # Convert the sorted list into a deque (queue structure)
    insertion_queue = deque(queue_items)

    return insertion_queue


def side_by_side_combine(
    workbook,
    template_sheets,
    supplier_sheets_dict,
    threshold=80,
    summary_option=False,
):
    """
    Combine the template sheets with the supplier sheets side by side.

    Parameters:
    workbook (openpyxl.Workbook): The workbook to combine the sheets into.
    template_sheets (list): A list of template sheets to combine.
    supplier_sheets_dict (dict): A dictionary mapping supplier names to their
        corresponding sheets to combine.
    threshold (int): The threshold for fuzzy matching between the template and
        supplier sheets.

    Returns:
    openpyxl.Workbook: The combined workbook with all the sheets.
    """
    st.toast(f"Combining files in progress...", icon="⏳")

    # Iterate over each template sheet
    for idx, template_sheet in enumerate(template_sheets):
        print(f"Processing template sheet: {template_sheet.title}")

        # Create a new sheet in the workbook for each template sheet
        target_sheet_template = workbook.create_sheet(
            f"{template_sheet.title} Template"[:30]
        )
        copy_sheet(template_sheet, target_sheet_template)

        # Create a new sheet in the workbook for side-by-side comparison
        target_sheet = workbook.create_sheet(f"Combined {template_sheet.title}"[:30])

        # Initialize variables to store column data and mismatched rows
        common_columns = (
            []
        )  # List of common column letters between the template and supplier sheets
        uncommon_columns = (
            []
        )  # List of columns that are not common between the template and supplier sheets
        mis_mat_rows_dict = (
            {}
        )  # Dictionary mapping supplier names to their corresponding mismatched rows
        supplier_value_columns_dict = (
            {}
        )  # Dictionary mapping supplier names to their corresponding columns of values
        supplier_colors = (
            {}
        )  # Dictionary mapping supplier names to their corresponding fill colors
        color_cycle = fill_color_switch()

        # Iterate over each supplier and process their sheet
        for supplier in supplier_sheets_dict:
            if supplier not in supplier_colors:
                supplier_colors[supplier] = next(color_cycle)

            supplier_sheet = supplier_sheets_dict[supplier][idx]
            com_columns, mis_mat_rows, supplier_value_columns = find_matching_cols(
                template_sheet, supplier_sheet, threshold
            )

            # Add matching columns to the final list (ensure no duplicates)
            common_columns.extend(
                [col for col in com_columns if col not in common_columns]
            )
            uncommon_columns.extend(
                [col for col in supplier_value_columns if col not in uncommon_columns]
            )

            # Store mismatched rows and value columns for the supplier
            mis_mat_rows_dict[supplier] = mis_mat_rows
            supplier_value_columns_dict[supplier] = supplier_value_columns

        # Copy common columns from template to target sheet
        queue = create_insertion_queue(common_columns, supplier_value_columns_dict)
        if not queue:
            print("No columns to process for this template sheet.")
            continue

        for i, item in enumerate(queue):
            col_letter = item["column_letter"]
            source = item["source"]
            header_fill_color = None
            mis_mat_rows = None

            # Determine the source sheet
            if source == "template":
                source_sheet = template_sheet
            else:
                source_sheet = supplier_sheets_dict[source][idx]
                header_fill_color = supplier_colors[source]
                mis_mat_rows = mis_mat_rows_dict[source]

            # Get column indices
            col_idx_source = column_index_from_string(col_letter)
            col_idx_target = i + 1  # Insert in the order of the queue

            # Copy column
            end_row_write = copy_column(
                source_sheet, target_sheet, col_idx_source, col_idx_target, mis_mat_rows
            )
            # format the header cell for the supplier if there are mismatched rows
            if header_fill_color:
                # header cell is the first bold cell in the column
                header_cell = target_sheet.cell(row=1, column=col_idx_target)
                header_cell.fill = PatternFill(
                    fill_type="solid", start_color=header_fill_color
                )
                # change the value of the header cell to include the supplier name
                if header_cell.value:
                    header_cell.value = f"{source}  {header_cell.value}"
                else:
                    header_cell.value = f"{source}"

                header_cell.font = Font(name="Arial", size=15, bold=True)
                header_cell.alignment = Alignment(
                    horizontal="center", vertical="center"
                )
                # apply color formatting to the data rows
                color_to_avoid = "FAA0A0"
                for row in range(2, end_row_write + 1):
                    cell_to_fill = target_sheet.cell(row=row, column=col_idx_target)
                    # only fill if the fill color is not FAA0A0
                    if (
                        cell_to_fill.fill
                        and cell_to_fill.fill.start_color
                        and cell_to_fill.fill.start_color.rgb
                    ):
                        existing_color = str(cell_to_fill.fill.start_color.rgb)
                    else:
                        existing_color = ""

                    # Fill only if the cell's color does not match the target color
                    if color_to_avoid not in existing_color:
                        cell_to_fill.fill = PatternFill(
                            fill_type="solid",
                            start_color=header_fill_color,
                            end_color=header_fill_color,
                        )
                # Add summary if requested
                if summary_option:
                    source_text = " "
                    for row in target_sheet.iter_rows(
                        min_row=2,
                        max_row=end_row_write,
                        min_col=col_idx_target,
                        max_col=col_idx_target,
                    ):
                        for cell in row:
                            # check if the cell is rich text
                            if cell.value is not None:
                                if isinstance(cell.value, CellRichText):
                                    source_text += " ".join(cell.value.as_list())
                                else:
                                    source_text += " " + str(cell.value) + " "
                    if len(source_text.split()) > 5:
                        summary = summarize_column_simple(source_text)
                        target_sheet.cell(
                            row=end_row_write + 1,
                            column=col_idx_target,
                            value="Summary:",
                        )
                        summary_cell = target_sheet.cell(
                            row=end_row_write + 2, column=col_idx_target
                        )
                        summary_cell.value = summary
                        summary_cell.font = Font(name="Arial", size=12, bold=False)
                        summary_cell.fill = PatternFill(
                            fill_type="solid", start_color="BFFFFF"
                        )
                        summary_cell.alignment = Alignment(
                            horizontal="center", vertical="center", wrap_text=True
                        )

        # Copy the template sheet attributes to the target sheet
        # copy_sheet_attributes(template_sheet, target_sheet)
        st.toast(f"{template_sheet.title} consolidated!", icon="✔️")
    # Remove the default sheet if it exists
    if "Sheet" in workbook.sheetnames:
        workbook.remove(workbook["Sheet"])

    # Final success toast
    st.toast("Side-By-Side File combined successfully! Ready to download", icon="🎉")

    return workbook


# Function to summarize text using Sumy
def summarize_column_simple(text: str, sentence_count: int = 3) -> str:
    """
    Summarize the given text using Sumy's Latent Semantic Analysis (LSA) summarizer.

    Args:
        text (str): The text to be summarized.
        sentence_count (int, optional): The number of sentences in the summary. Defaults to 3.

    Returns:
        str: The summary of the text.
    """
    try:
        # Create a Sumy parser from the text
        parser = PlaintextParser.from_string(text, Tokenizer("english"))

        # Create a Sumy LSA summarizer
        summarizer = LsaSummarizer()

        # Set the stop words for the summarizer to English
        summarizer.stop_words = "english"

        # Summarize the text
        summary = summarizer(parser.document, sentence_count)

        # Join the summary sentences into a single string
        return " ".join(str(sentence) for sentence in summary)
    except Exception as e:
        # Return an error message if there's an exception
        return f"Error summarizing text: {e}"


# function to save the consolidated file and let it be downloadable by the user
def save_consolidated_file(consolidated):
    """
    Save the consolidated file to a BytesIO object and return it for downloading.

    Args:
        consolidated (openpyxl.Workbook): The consolidated workbook.

    Returns:
        io.BytesIO: The BytesIO object containing the workbook.
    """
    # Create a BytesIO object to store the workbook
    file_stream = io.BytesIO()

    # Save the workbook to the BytesIO object
    consolidated.save(file_stream)

    # Reset the file pointer to the beginning of the BytesIO object
    file_stream.seek(0)

    return file_stream


def append_logo(workbook, image_path, image_scale=0.8):
    """
    Append the logo to each sheet in the workbook.

    Args:
        workbook (openpyxl.Workbook): The workbook to append the logo to.
        image_path (str): The path to the image file. If None or empty, no image is appended.
        image_scale (float): The scale of the image. Default is 0.8.
    """
    # get all the sheet names
    sheet_names = workbook.sheetnames
    # get all the worksheets
    worksheets = [workbook[sheet_name] for sheet_name in sheet_names]
    for worksheet in worksheets:
        # add logo to the first cell of the worksheet
        if image_path:
            # import the Image class from openpyxl.drawing.image
            from openpyxl.drawing.image import Image

            # create an Image object
            logo = Image(image_path)
            # scale the image
            logo.width = int(logo.width * image_scale)
            logo.height = int(logo.height * image_scale)
            # add the image to the worksheet
            worksheet.add_image(logo, "A1")

    return workbook


def copy_column(
    source_sheet, target_sheet, source_col_idx, target_col_idx, mis_mat_rows=None
):
    """
    Copy a column from the source sheet to the target sheet.

    Args:
        source_sheet (openpyxl.Worksheet): The source sheet.
        target_sheet (openpyxl.Worksheet): The target sheet.
        source_col_idx (int): The column index of the source sheet.
        target_col_idx (int): The column index of the target sheet.
        mis_mat_rows (list): The list of rows that have mismatched values between the source and target sheets.
    Returns:
        int: The row index of the last row copied.
    """
    source_col_letter = get_column_letter(source_col_idx)
    target_col_letter = get_column_letter(target_col_idx)

    # Initialize target_row at the first row of the target sheet
    hidden_count = 0
    end_row_idx = 500  # default value
    # Track merged cells in the source column
    # quick check the first 60 rows to see if there are any values, if all is empty, break the loop
    max_rows_to_check = min(source_sheet.max_row, 60)
    check_empty = 0
    for row in range(1, max_rows_to_check + 1):
        source_cell = source_sheet.cell(row=row, column=source_col_idx)
        if source_cell.value not in (
            None,
            "",
        ):  # If a non-empty cell is found, exit early
            break
        check_empty += 1

    # If all rows are empty, exit the function
    if check_empty == max_rows_to_check:
        return end_row_idx
    target_row = 1
    merged_dict = generate_merged_dict(source_sheet)
    # Iterate over each row in the source sheet
    empty_rows_cont = 0
    if mis_mat_rows:
        # covert to a list of tuples for easy comparison
        mis_mat_rows = [coordinate_from_string(row) for row in mis_mat_rows]
    for row in range(1, min(source_sheet.max_row + 1, 500)):
        source_cell = source_sheet.cell(row=row, column=source_col_idx)
        target_cell = target_sheet.cell(row=target_row, column=target_col_idx)
        if source_cell.value is None or source_cell.value == "":
            empty_rows_cont += 1
            if empty_rows_cont > 80:
                end_row_idx = row - empty_rows_cont + 1

                break
        else:
            empty_rows_cont = 0
        # Skip hidden rows in the source sheet
        if getattr(source_sheet.row_dimensions[row], "hidden", False):
            hidden_count += 1
            if hidden_count > 60:

                end_row_idx = row - empty_rows_cont + 1

                break
        else:
            hidden_count = 0

        source_cell_value = copy(source_cell.value)
        target_cell.value = source_cell_value
        target_cell.data_type = copy(source_cell.data_type)

        try:
            if source_cell.has_style:
                target_cell.font = copy(source_cell.font)
                target_cell.border = copy(source_cell.border)
                target_cell.fill = copy(source_cell.fill)
                target_cell.number_format = copy(source_cell.number_format)
                target_cell.protection = copy(source_cell.protection)
                target_cell.alignment = copy(source_cell.alignment)
        except Exception as e:
            print(f"Error copying styles for cell {source_cell.coordinate}: {e}")

        if source_cell.hyperlink:
            target_cell.hyperlink = source_cell.hyperlink

        if source_cell.comment:
            target_cell.comment = copy(source_cell.comment)
        if mis_mat_rows:
            col_s, row_s = coordinate_from_string(source_cell.coordinate)
            for mis_row in mis_mat_rows:
                if mis_row[1] == row_s:
                    # conver mis_row[0] to column index
                    col_mis_idx = column_index_from_string(mis_row[0])
                    if (
                        source_col_idx - col_mis_idx == 1
                    ):  # meaning the mismatched column is the previous column
                        target_cell.fill = PatternFill(
                            start_color="FAA0A0", end_color="FAA0A0", fill_type="solid"
                        )

        target_row += 1

    if end_row_idx == 500:
        end_row_idx = target_row - empty_rows_cont + 1

    # Perform merging of cells after copying data
    merge_columns_in_target_sheet(
        target_sheet, merged_dict, source_col_idx, target_col_idx
    )

    # Copy column width and hidden property
    source_dim = source_sheet.column_dimensions[source_col_letter]
    target_dim = target_sheet.column_dimensions[target_col_letter]
    if hasattr(source_dim, "width"):
        target_dim.width = source_dim.width
    if hasattr(source_dim, "hidden"):
        target_dim.hidden = source_dim.hidden

    return end_row_idx


# function to copy the sheet from source to target
def copy_sheet(source_sheet, target_sheet):
    """
    Copy all columns and sheet-level attributes from the source sheet to the target sheet.

    Args:
        source_sheet (openpyxl.Worksheet): The sheet to copy from.
        target_sheet (openpyxl.Worksheet): The sheet to copy to.
    """
    hidden_count = 0  # Counter to track consecutive hidden columns
    for col in source_sheet.iter_cols(max_col=min(source_sheet.max_column, 100)):
        col_idx = col[0].column
        col_letter = get_column_letter(col_idx)

        # Check if the column is hidden
        if source_sheet.column_dimensions[col_letter].hidden:
            hidden_count += 1
            if (
                hidden_count > 60
            ):  # Stop copying if more than 60 consecutive columns are hidden
                break
            continue

        # Copy the column to the same index in the target
        _ = copy_column(source_sheet, target_sheet, col_idx, col_idx)

    # Copy sheet-level attributes
    copy_sheet_attributes(source_sheet, target_sheet)


def copy_sheet_attributes(source_sheet, target_sheet):
    """
    Copy various attributes from a source sheet to a target sheet.

    Args:
        source_sheet (openpyxl.Worksheet): The sheet to copy attributes from.
        target_sheet (openpyxl.Worksheet): The sheet to copy attributes to.
    """
    # Copy basic sheet properties
    target_sheet.sheet_format = copy(source_sheet.sheet_format)
    target_sheet.sheet_properties = copy(source_sheet.sheet_properties)
    target_sheet.page_margins = copy(source_sheet.page_margins)
    target_sheet.page_setup = copy(source_sheet.page_setup)
    target_sheet.print_options = copy(source_sheet.print_options)
    target_sheet.auto_filter = copy(source_sheet.auto_filter)
    target_sheet.print_area = source_sheet.print_area
    target_sheet.freeze_panes = source_sheet.freeze_panes

    # Copy the row heights
    for rn, source_row in source_sheet.row_dimensions.items():
        target_sheet.row_dimensions[rn].ht = copy(source_row.ht)

    # Copy merged cell ranges
    for merged_range in source_sheet.merged_cells.ranges:
        target_sheet.merge_cells(str(merged_range))

    # Copy sheet protection settings
    target_sheet.protection = copy(source_sheet.protection)


def get_files(supplier_info, sheet_indexes, doc_type):
    """
    Read the specified files for each supplier and return the DataFrames and sheets.

    Args:
        supplier_info (list): List of dictionaries containing supplier information.
        sheet_indexes (list): List of sheet indexes to read (0-indexed).
        doc_type (str): Document type to read (either "RFP" or "Proposal").

    Returns:
        tuple: Two dictionaries, the first containing the DataFrames, the second containing the sheets.
    """
    dfs_dict = {}
    worksheets_dict = {}
    with st.spinner("Reading files..."):
        for supplier in supplier_info:
            if supplier.get(doc_type):
                dfs_dict[supplier["name"]] = []
                worksheets_dict[supplier["name"]] = []
            else:
                # Warn if no file was found
                st.warning(
                    f"No {doc_type} file found for supplier {supplier['name']}.",
                    icon="⚠️",
                )
                continue
            # Read the file
            sup_excel = load_workbook(
                supplier[doc_type],
                rich_text=st.session_state.richtext_option,
                data_only=True,
            )
            # Get the visible sheets
            sup_sheets = [
                sup_excel[sheet]
                for sheet in sup_excel.sheetnames
                if sup_excel[sheet].sheet_state == "visible"
            ]
            # Iterate over the specified sheet indexes
            for sheet_idx in sheet_indexes:
                sheet = sup_sheets[sheet_idx]
                # Add the sheet to the dictionary
                worksheets_dict[supplier["name"]].append(sheet)

    # Show a toast when done
    st.toast("Supplier Files read successfully! 📚", icon="✅")

    return dfs_dict, worksheets_dict


def write_summary_to_sheet(summary_df, grand_total_df, summary_sheet):
    """
    Write the summary DataFrame and grand total DataFrame to the summary sheet.

    Args:
        summary_df (pd.DataFrame): The summary DataFrame
        grand_total_df (pd.DataFrame): The grand total DataFrame
        summary_sheet (openpyxl.worksheet.worksheet.Worksheet): The summary sheet

    Note:
        This function writes the summary DataFrame to the sheet, then writes the
        grand total DataFrame to the sheet below the summary DataFrame. It also
        formats the sheet by applying borders to all cells and setting the column
        widths to auto based on the content.
    """
    if not summary_df.empty:
        # Pivot the summary DataFrame
        merged_df = summary_df.pivot_table(
            index=["Category", "Subcategory"],
            values=[
                col
                for col in summary_df.columns
                if col not in ["Category", "Subcategory"]
            ],
            aggfunc="first",
        ).reset_index()

        # Write merged_df to the sheet
        for r in dataframe_to_rows(merged_df, index=False, header=True):
            summary_sheet.append(r)

        # Separate `merged_df` and `grand_total_df` visually in the sheet
        summary_sheet.append([])
        summary_sheet.append(
            ["Grand Total Summary"]
        )  # Add a title row for grand_total_df

        # Write grand_total_df to the sheet
        for r in dataframe_to_rows(grand_total_df, index=False, header=True):
            summary_sheet.append(r)

        # Merge cells with the same value in column A (Category column)
        current_value = None
        start_row = None
        for row in range(2, summary_sheet.max_row + 1):  # Skip header row
            cell_value = summary_sheet.cell(row=row, column=1).value
            if cell_value != current_value:
                if start_row and current_value is not None:
                    summary_sheet.merge_cells(
                        start_row=start_row,
                        start_column=1,
                        end_row=row - 1,
                        end_column=1,
                    )
                current_value = cell_value
                start_row = row
        if start_row and current_value is not None:
            summary_sheet.merge_cells(
                start_row=start_row,
                start_column=1,
                end_row=summary_sheet.max_row,
                end_column=1,
            )

        # Format the sheet
        header_fill = PatternFill(
            start_color="4472C4", end_color="4472C4", fill_type="solid"
        )
        font_style_header = Font(b=True, color="FFFFFF")
        border_style = Side(border_style="thin", color="000000")

        # Apply borders and styles to all cells
        grand_total_row = 1
        for row in summary_sheet.iter_rows(
            min_row=1,
            max_row=summary_sheet.max_row,
            min_col=1,
            max_col=summary_sheet.max_column,
        ):
            for cell in row:
                cell.border = Border(
                    left=border_style,
                    right=border_style,
                    top=border_style,
                    bottom=border_style,
                )
                if cell.row == 1:
                    cell.fill = header_fill
                    cell.font = font_style_header
                elif cell.value == "Grand Total Summary":
                    # fill the cell with a different color
                    cell.fill = PatternFill(
                        start_color="FF0000", end_color="FF0000", fill_type="solid"
                    )
                    cell.font = font_style_header
                    grand_total_row = cell.row

        # Set the column width to auto based on the content
        for col in range(1, summary_sheet.max_column + 1):
            column = get_column_letter(col)
            max_length = max(
                len(str(summary_sheet.cell(row=row, column=col).value) or "")
                for row in range(1, summary_sheet.max_row + 1)
            )
            summary_sheet.column_dimensions[column].width = max_length + 2

        # Plot bar chart of the grand total summary
        chart = BarChart()
        chart.type = "col"  # Column chart
        chart.grouping = "clustered"  # Set grouping to clustered
        chart.title = "Supplier Price Comparison by Category"
        chart.x_axis.title = "Category"
        chart.y_axis.title = "Price"

        data = Reference(
            summary_sheet,
            min_col=2,
            min_row=grand_total_row + 1,
            max_row=summary_sheet.max_row,
            max_col=summary_sheet.max_column - 1,
        )
        categories = Reference(
            summary_sheet,
            min_col=1,
            min_row=grand_total_row + 2,
            max_row=summary_sheet.max_row,
        )
        chart.add_data(data, titles_from_data=True)
        chart.set_categories(categories)
        summary_sheet.add_chart(chart, "A" + str(summary_sheet.max_row + 2))
        # Display the category in the x-axis

        # Add a title for the chart
        summary_sheet.cell(
            row=summary_sheet.max_row + 4,
            column=1,
            value="Supplier Price Comparison by Category",
        )


def create_summary_price_table(summary_sheet, price_sheet, supplier_names):
    """
    Create a summary price table by extracting price data from the price sheet
    and writing it to the summary sheet.

    Args:
        summary_sheet (openpyxl.Worksheet): The worksheet to write the summary data.
        price_sheet (openpyxl.Worksheet): The worksheet to extract price data from.
        supplier_names (list): A list of supplier names to map columns.

    Returns:
        int: 1 if the summary table is created successfully, 0 otherwise.
    """
    # Extract headers and map columns to suppliers
    headers_dict = {}
    for col in price_sheet.iter_cols():
        col_letter = get_column_letter(col[0].column)
        headers = [(cell.value, cell.row) for cell in col if cell.font.bold]
        headers_dict[col_letter] = headers

    # Identify columns for price data
    price_label_col = None
    max_len = 0
    supplier_cols_dict = defaultdict(list)
    for key, value in headers_dict.items():
        if value and value[0][1] == 1 and value[0][0] in supplier_names:
            supplier_cols_dict[value[0][0]].append(key)
        else:
            if len(value) > max_len:
                max_len = len(value)
                price_label_col = key

    # Compile summary data
    summary_data = []
    for supplier, cols in supplier_cols_dict.items():
        for col in cols:
            col_headers = headers_dict.get(col, [])
            if len(col_headers) < 2:
                continue
            category = col_headers[1][0]
            upper_row = headers_dict[price_label_col][0][1]
            lower_row = headers_dict[price_label_col][0][1]
            for cate, row in headers_dict[price_label_col]:
                if row > lower_row:
                    lower_row = row
                value = price_sheet[f"{col}{row}"].value
                price_value = re.sub(r"[^\d.]", "", str(value))
                if price_value:
                    try:
                        price_value = round(float(price_value), 2)
                    except ValueError:
                        continue
                    summary_data.append(
                        {
                            "Category": category,
                            "Subcategory": cate,
                            supplier: price_value,
                        }
                    )
                else:
                    total_price = 0
                    for row in range(upper_row, lower_row + 1):
                        value = price_sheet[f"{col}{row}"].value
                        price_value = re.sub(r"[^\d.]", "", str(value))
                        if price_value:
                            try:
                                price_value = float(price_value)
                            except ValueError:
                                continue
                            total_price += price_value
                    if total_price > 0:
                        total_price = round(total_price, 2)
                        summary_data.append(
                            {
                                "Category": category,
                                "Subcategory": cate,
                                supplier: total_price,
                            }
                        )

    # Convert summary data into DataFrame
    summary_df = pd.DataFrame(summary_data)

    # Pivot the DataFrame and reset the index
    if not summary_df.empty:
        merged_df = summary_df.pivot_table(
            index=["Category", "Subcategory"],
            values=[
                col
                for col in summary_df.columns
                if col not in ["Category", "Subcategory"]
            ],
            aggfunc="first",
        ).reset_index()

        # Ensure that only existing suppliers are used
        existing_suppliers = [
            supplier for supplier in supplier_names if supplier in merged_df.columns
        ]
        if not existing_suppliers:
            print("No valid suppliers found in the DataFrame columns.")
            return summary_sheet
        try:
            grand_total_rows = merged_df[
                merged_df["Subcategory"].str.contains(
                    "grand total", case=False, na=False
                )
            ].index

            if grand_total_rows.size > 0:
                # Extract the rows containing 'Grand Total'
                grand_total_df = merged_df.loc[grand_total_rows]
                # drop the 'Subcategory' column
                grand_total_df = grand_total_df.drop(columns=["Subcategory"])
            else:
                # Aggregate the data to compute the grand total
                numeric_columns = merged_df.select_dtypes(include="number").columns
                grand_total_df = (
                    merged_df.groupby("Category")[numeric_columns].sum().reset_index()
                )
        except Exception as e:
            print(f"Error Creating Summary Table: {e}")
            return 0
    else:
        print("The summary_df is empty; ensure your input data is correct.")
        return 0

    # Write the summary and grand total data to the summary sheet
    write_summary_to_sheet(merged_df, grand_total_df, summary_sheet)
    return 1


# streamlit_app\kellanova_logo.png
st.image(r"assets/kellanova_logo.png", width=200)

# Check if suppliers are set up
if "suppliers" not in st.session_state or len(st.session_state.suppliers) == 0:
    st.error("No supplier data found. Please complete the setup first.")
    st.stop()

# Check if document types are configured
if "doc_types" not in st.session_state or len(st.session_state.doc_types) == 0:
    st.error("Document types not configured. Please complete the setup first.")
    st.stop()

if "event_name" not in st.session_state or "event_option" not in st.session_state:
    st.error("Event name or option not configured. Please complete the setup first.")
    st.stop()

# Check if each supplier has the necessary files and names
for i, supplier in enumerate(st.session_state.suppliers):
    if (
        not supplier.get("name")
        or not supplier.get(st.session_state.doc_types[0])
        or not supplier.get(st.session_state.doc_types[1])
    ):
        st.error(
            f"Supplier {i+1} is missing required data (name or template files). Please complete the setup first."
        )
        st.stop()

### Session State Variables Retrieval

event_name = st.session_state.event_name
event_option = st.session_state.event_option
# logo path
st.session_state.logo_path = r"assets/kellanova_logo.png"
doc_type1, doc_type2 = st.session_state.doc_types


st.write("# RFP Files Consolidator")
suppliers_html = ""
color_cycle = fill_color_switch()

for i, supplier in enumerate(st.session_state.suppliers):
    supplier_name = supplier["name"]
    supplier_color = next(color_cycle)

    suppliers_html += f'<span style="font-weight: bold; color: #{supplier_color}; margin-right: 10px;">{supplier_name}</span>'

# Render the formatted text with bold event and option names
st.markdown(
    f"""
    <div>
        <p style="font-style: italic;">Current Event: <span style="font-weight: bold;">{st.session_state.event_name}</span></p>
        <p style="font-style: italic;">Document Configuration: <span style="font-weight: bold;">{st.session_state.event_option}</span></p>
        <p style="font-style: italic;">Participants: {suppliers_html}</p>
    </div>
    """,
    unsafe_allow_html=True,
)

### Toggle between Rich Text and Plain Text, Rich Text will retain all original formatting but may generate a warning message when opening the consolidated file. By uncommenting the line below, you can toggle between Rich Text and Plain Text. right now, Rich Text is set to True by default.

# richtext_option = st.checkbox(
#     "Rich Text",
#     value=False,
#     key="richtext_option",
#     help="Keep all the original formatting within a cell when consolidating",
# )

st.session_state.richtext_option = True

### Pricing Sheets Consolidation

st.markdown("### :green[For **Pricing**]")
st.markdown(
    "***Note:** Summary tab and chart for Pricing only works in **Side by Side** consolidation.*"
)

template_pri = st.session_state.template_files[doc_type1]
wb_template_pri = load_workbook(
    template_pri, rich_text=st.session_state.richtext_option, data_only=True
)
all_sheets_pri = wb_template_pri.sheetnames

pricing_sheets_list = st.multiselect(
    "Please select Pricing sheet(s) to consolidate", all_sheets_pri, all_sheets_pri
)

chosen_sheets_pri_idx = [all_sheets_pri.index(sheet) for sheet in pricing_sheets_list]

if "pri_comb_mode" not in st.session_state:
    st.session_state.pri_comb_mode = "Side by Side"


st.radio(
    "Please select consolidation method:",
    ("Side by Side", "Separate Sheets"),
    index=0,
    key="pri_comb_mode",
)


if st.button("Consolidate", key="consolidate_pri"):
    consolidated_pri = openpyxl.Workbook()
    consolidated_pri.remove(consolidated_pri.active)
    dfs_pri_dict, sheets_pri_dict = get_files(
        st.session_state.suppliers, chosen_sheets_pri_idx, doc_type1
    )
    template_sheets_pri = [wb_template_pri[sheet] for sheet in pricing_sheets_list]
    with st.spinner("Processing... Please wait."):
        if st.session_state.pri_comb_mode == "Side by Side":
            consolidated_pri = side_by_side_combine(
                consolidated_pri, template_sheets_pri, sheets_pri_dict
            )
            supplier_names = list(sheets_pri_dict.keys())
            # iterate over the sheets in consolidated_pri and create a summary sheet
            for sheet in consolidated_pri.worksheets:
                # Skip template or non-price sheets if needed
                if "Combined" in sheet.title:
                    # Create a new summary sheet
                    summary_sheet = consolidated_pri.create_sheet(
                        title=f"Summary of {sheet.title}"[:30]
                    )
                    status_sum = create_summary_price_table(
                        summary_sheet, sheet, supplier_names
                    )
                    # Move the summary sheet to the leftmost position
                    if status_sum:
                        consolidated_pri._sheets.remove(summary_sheet)
                        consolidated_pri._sheets.insert(0, summary_sheet)
                    else:
                        # remove the sheet if the summary is not created
                        consolidated_pri.remove(summary_sheet)

        elif st.session_state.pri_comb_mode == "Separate Sheets":
            consolidated_pri = separate_sheet_combine(
                consolidated_pri, template_sheets_pri, sheets_pri_dict
            )
        consolidated_pri = append_logo(consolidated_pri, st.session_state.logo_path)
        file_stream_p = save_consolidated_file(consolidated_pri)
    if file_stream_p is None:
        st.error("Failed to save the consolidated file. Please try again.")
        st.stop()
    # save to session state
    st.session_state.consolidated_p = file_stream_p
    st.success("Pricing sheets consolidated successfully!", icon="✅")


if st.session_state.get("consolidated_p"):
    st.download_button(
        f"💾 Download {event_name}_{doc_type1}_consolidated.xlsx",
        data=st.session_state.consolidated_p,  # Changed from consolidated_q
        file_name=f"{event_name}_{doc_type1}_consolidated.xlsx",
    )
else:
    st.session_state.consolidated_p = None


### Questionnaire Sheets Consolidation

st.markdown("#### :orange[For **Questionnaire**]")
template_ques = st.session_state.template_files[doc_type2]
wb_template_ques = load_workbook(
    template_ques, rich_text=st.session_state.richtext_option, data_only=True
)
all_sheets_ques = wb_template_ques.sheetnames

questionnaire_sheets_list = st.multiselect(
    "Please select Questionnaire sheet(s) to consolidate",
    all_sheets_ques,
    all_sheets_ques,
)

chosen_sheets_ques_idx = [
    all_sheets_ques.index(sheet) for sheet in questionnaire_sheets_list
]

if "ques_comb_mode" not in st.session_state:
    st.session_state.ques_comb_mode = "Side by Side"

st.radio(
    "Please select consolidation method:",
    ("Side by Side", "Separate Sheets"),
    index=0,
    key="ques_comb_mode",
)


# add a summary option tickbox
summary_option = st.checkbox(
    "Questionnaire summary included",
    value=False,
    key="summary_option",
    help="Include a summary of the supplier responses at the end of each column. Side by Side mode only.",
)


if st.button("Consolidate", key="consolidate_ques"):
    consolidated_ques = openpyxl.Workbook()
    consolidated_ques.remove(consolidated_ques.active)
    dfs_ques_dict, sheets_ques_dict = get_files(
        st.session_state.suppliers, chosen_sheets_ques_idx, doc_type2
    )
    template_sheets_ques = [
        wb_template_ques[sheet] for sheet in questionnaire_sheets_list
    ]
    with st.spinner("Processing... Please wait."):
        if st.session_state.ques_comb_mode == "Side by Side":
            consolidated_ques = side_by_side_combine(
                consolidated_ques,
                template_sheets_ques,
                sheets_ques_dict,
                summary_option=st.session_state.summary_option,
            )
        elif st.session_state.ques_comb_mode == "Separate Sheets":
            consolidated_ques = separate_sheet_combine(
                consolidated_ques, template_sheets_ques, sheets_ques_dict
            )
        consolidated_ques = append_logo(consolidated_ques, st.session_state.logo_path)
        file_stream_q = save_consolidated_file(consolidated_ques)

    # save to session state
    st.session_state.consolidated_q = file_stream_q
    st.success("Questionnaire sheets consolidated successfully!", icon="✅")

download_questionnaire = False

if st.session_state.get("consolidated_q"):
    st.download_button(
        f"💾 Download {event_name}_{doc_type2}_consolidated.xlsx",
        data=st.session_state.consolidated_q,
        file_name=f"{event_name}_{doc_type2}_consolidated.xlsx",
    )
    download_questionnaire = True
else:
    st.session_state.consolidated_q = None
