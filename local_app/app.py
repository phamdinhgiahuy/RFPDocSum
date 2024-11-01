from flask import Flask, redirect, url_for, render_template, request, session, flash
from datetime import timedelta
from flask_sqlalchemy import SQLAlchemy
from second import second
import os
import glob
import re
from openpyxl import load_workbook
import pandas as pd
from consolidate import copy_sheet
import openpyxl
from flask import jsonify

app = Flask(__name__)
app.secret_key = "s3cr3t"
app.permanent_session_lifetime = timedelta(minutes=15)
app.register_blueprint(second, url_prefix="/admin")


app.config["SQLALCHEMY_DATABASE_URI"] = "sqlite:///users.sqlite3"
app.config["SQLALCHEMY_TRACK_MODIFICATIONS"] = False

db = SQLAlchemy(app)


class Users(db.Model):
    _id = db.Column("id", db.Integer, primary_key=True)
    name = db.Column(db.String(100))
    email = db.Column(db.String(100))

    def __init__(self, name, email):
        self.name = name
        self.email = email


@app.route("/")
@app.route("/home/")
def home():
    return render_template("index.html")


@app.route("/view")
def view():
    return render_template("view.html", values=Users.query.all())


@app.route("/admin/")
def admin():
    return redirect(url_for("user", name="Admin"))


@app.route("/login/", methods=["GET", "POST"])
def login():
    if request.method == "POST":
        session.permanent = True
        user = request.form["username"]
        session["user"] = user

        found_user = Users.query.filter_by(name=user).first()
        if found_user:
            session["email"] = found_user.email

        else:
            usr = Users(user, "")
            db.session.add(usr)
            db.session.commit()

        flash(f"You are now logged in, {user}!", "success")
        return redirect(url_for("user"))
    else:
        if "user" in session:
            return redirect(url_for("user"))
        return render_template("login.html")


@app.route("/user/", methods=["GET", "POST"])
def user():
    email = None
    if "user" in session:
        usr = session["user"]
        if request.method == "POST":
            email = request.form["email"]
            session["email"] = email
            found_user = Users.query.filter_by(name=usr).first()
            found_user.email = email
            db.session.commit()
            flash("Email was saved!", "success")
        else:
            if "email" in session:
                email = session["email"]
        return render_template("user.html", user=usr, email=email)
    else:
        flash("You are not logged in!", "danger")
        return redirect(url_for("login"))


@app.route("/delete/<int:id>", methods=["POST"])
def delete(id):
    user_to_delete = Users.query.get_or_404(id)
    try:
        db.session.delete(user_to_delete)
        db.session.commit()
        flash(f"User {user_to_delete.name} was successfully deleted.", "success")
    except:
        db.session.rollback()
        flash("There was a problem deleting the user.", "danger")

    return redirect(url_for("view"))


@app.route("/logout")
def logout():
    if "user" in session:
        usr = session["user"]
        flash(f"You have been logged out, {usr}!", "info")
    session.pop("user", None)
    session.pop("email", None)
    session.pop("price_labels_set", None)
    return redirect(url_for("login"))


@app.route("/input", methods=["GET", "POST"])
def input_rfp():
    if "user" not in session:
        flash("You are not logged in!", "danger")
        return redirect(url_for("login"))
    price_labels_set = set()
    if request.method == "POST":
        # Retrieve form data
        event_name = request.form.get("event_name")
        num_suppliers = int(request.form.get("num_suppliers"))
        suppliers = [
            request.form.get(f"supplier_{i}") for i in range(1, num_suppliers + 1)
        ]
        flash(
            f"Sucessfully configured the RFP Event {event_name} for these participating Suppliers: {suppliers} Please wait for the consolidation process to complete.",
            "warning",
        )
        # Validate form data
        # Retrieve the uploaded file
        uploaded_file = request.files.get("rfp_file")

        if uploaded_file:
            # Get the file path and extract the parent folder
            filename = uploaded_file.filename
            folder_path = os.path.dirname(filename)

        files = []
        for file in glob.glob(
            "../Project Files/Datasets/KPS+/**/*.xlsx", recursive=True
        ):
            if "Pricing" in file:
                files.append(file)
            # Matching suppliers to files
        supplier_file_mapping = {}
        for supplier in suppliers:
            matched_files = [file for file in files if supplier.lower() in file.lower()]
            if matched_files:
                supplier_file_mapping[supplier] = matched_files
        # Flash the result
        for supplier in suppliers:
            if supplier in supplier_file_mapping:
                flash(
                    f"Supplier {supplier} found in the following files: {supplier_file_mapping[supplier]}.",
                    "success",
                )
            else:
                flash(
                    f"Supplier {supplier} was not found in any of the files. Please check the spelling and try again.",
                    "danger",
                )
        data_price = {}
        data_imple = {}
        pricing_sheets = {}
        implement_sheets = {}
        for file in files[1:]:
            supplier = load_workbook(file)
            for sheet in supplier.sheetnames:
                if "Pricing" in sheet:
                    df_temp = pd.read_excel(
                        file,
                        sheet_name=sheet,
                        header=7,
                    )
                    key_price = file.split("/")[-1] + "_" + "price"
                    data_price[key_price] = (
                        df_temp  # .loc[:, ~df_temp.columns.str.contains('^Unnamed')].dropna(how='all')
                    )
                    pricing_sheets[file.split("/")[-1].split("-")[-1]] = supplier[sheet]
                elif "Implementation" in sheet:
                    df_temp = pd.read_excel(file, sheet_name=sheet, header=4)
                    key_imple = file.split("/")[-1] + "_" + "implement"
                    data_imple[key_imple] = df_temp
                    implement_sheets[file.split("/")[-1].split("-")[-1]] = supplier[
                        sheet
                    ]
        keys_price = list(data_price.keys())
        keys_imple = list(data_imple.keys())
        price_labels = []
        for key in keys_price:
            df_test = data_price[key]
            # rename the first col to be "Deliverable"
            df_test = df_test.rename(columns={df_test.columns[0]: "Phase"})
            # drop unnamed columns
            df_test = df_test.loc[:, ~df_test.columns.str.contains("^Unnamed")].dropna(
                how="all"
            )
            # print(df_test.columns)
            # clean the Deliverable col
            # find all the numbers and '-' in the Deliverable column
            df_test["Deliverable_clean"] = df_test["Deliverable"].apply(
                lambda x: ("".join(x)).split("-")[-1] if type(x) == list else x
            )  # str.findall(r"\d+|-|,|\.")
            # join all the numbers and '-' in the Deliverable_clean column if not NaN
            df_test["Deliverable_clean"] = df_test["Deliverable_clean"].apply(
                lambda x: ("".join(x)).split("-")[-1] if type(x) == list else x
            )
            mils = df_test["Milestone"].values
            phases = df_test["Phase"].values
            for i in range(len(mils)):
                com_text = f"{mils[i]} - {phases[i]}"
                com_text = com_text.replace("nan", "").strip()
                price_labels.append(com_text)
        price_labels_set = set(price_labels)

        consolidated = openpyxl.Workbook()
        for key, sheet in pricing_sheets.items():
            target_sheet = consolidated.create_sheet(sheet.title)
            copy_sheet(sheet, target_sheet)

        for key, sheet in implement_sheets.items():
            target_sheet = consolidated.create_sheet(sheet.title)
            copy_sheet(sheet, target_sheet)
        if "Sheet" in consolidated.sheetnames:  # remove default sheet
            consolidated.remove(consolidated["Sheet"])
        consolidated.save("consolidated.xlsx")
        flash(
            f"Finished! Consolidated consolidated.xlsx file created at {os.getcwd()}!",
            "success",
        )
        flash(
            f"Please select the price labels that you would like to use for the RFP Event {event_name} from the list below {price_labels_set}.",
            "info",
        )

        # Redirect after POST to clear form data
        return redirect(url_for("input_rfp"))

    return render_template("input.html")


@app.route("/get_price_labels")
def get_price_labels():
    # Retrieve price_labels_set from session
    # if not logged in, return empty list
    if "user" not in session:
        # remove the price_labels_set from the session
        session.pop("price_labels_set", None)
        return jsonify([])
    price_labels_set = session.get("price_labels_set", [])
    return jsonify(price_labels_set)


if __name__ == "__main__":
    with app.app_context():
        # Code that accesses the db object goes here
        db.create_all()
    app.run(debug=True)
